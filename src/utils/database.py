"""
Módulo de gerenciamento de banco de dados JSON (Legado)

Este módulo implementa a classe DatabaseManager que gerencia o armazenamento
de dados em arquivos JSON para o sistema RH Control. É considerado legado e
será eventualmente substituído pelo database_sql.py, mas mantém compatibilidade
durante o período de transição.

Autor: Sistema RH Control
Data: 2024
Versão: 1.0 (Legado)
"""

import json
import os
from datetime import datetime
from typing import List, Dict, Optional, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


# ============================================================================
# CLASSE PRINCIPAL - DATABASE MANAGER
# ============================================================================

class DatabaseManager:
    """
    Gerenciador de banco de dados JSON legado.
    
    Esta classe gerencia o armazenamento de dados em arquivos JSON para
    funcionários, afastamentos e usuários. Implementa operações CRUD completas
    e mantém compatibilidade com a interface do DatabaseSQL para facilitar
    a migração futura.
    
    Attributes:
        data_dir (str): Diretório onde os arquivos JSON são armazenados
        employees_file (str): Caminho do arquivo de funcionários
        afastamentos_file (str): Caminho do arquivo de afastamentos (compatibilidade)
        users_file (str): Caminho do arquivo de usuários
    """
    
    def __init__(self, data_dir: str = "data"):
        """
        Inicializa o gerenciador de banco de dados JSON.
        
        Args:
            data_dir: Diretório para armazenamento dos arquivos JSON
        """
        self.data_dir = data_dir
        self.employees_file = os.path.join(data_dir, "employees.json")
        self.afastamentos_file = os.path.join(data_dir, "absences.json") # Renomeado para compatibilidade com o código do usuário
        self.users_file = os.path.join(data_dir, "users.json")
        
        # Criar diretório de dados se não existir
        self._ensure_data_directory()
        
        # Inicializar arquivos JSON se não existirem
        self._initialize_files()
    
    # ========================================================================
    # MÉTODOS DE INICIALIZAÇÃO
    # ========================================================================
    
    def _ensure_data_directory(self) -> None:
        """Cria o diretório de dados se não existir."""
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
    
    def _initialize_files(self) -> None:
        """Inicializa os arquivos JSON com estruturas vazias se não existirem."""
        if not os.path.exists(self.employees_file):
            self._save_json(self.employees_file, [])
        
        if not os.path.exists(self.afastamentos_file):
            self._save_json(self.afastamentos_file, [])
        
        if not os.path.exists(self.users_file):
            self._save_json(self.users_file, [])
    
    # ========================================================================
    # MÉTODOS AUXILIARES DE ARQUIVO (PRIVADOS PARA COMPATIBILIDADE)
    # ========================================================================
    
    def _load_json(self, filepath: str) -> List[Dict[str, Any]]:
        """
        Carrega dados de um arquivo JSON.
        
        Args:
            filepath: Caminho do arquivo JSON
            
        Returns:
            Lista de dicionários com os dados carregados
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            return []
    
    def _save_json(self, filepath: str, data: List[Dict[str, Any]]) -> None:
        """
        Salva dados em um arquivo JSON.
        
        Args:
            filepath: Caminho do arquivo JSON
            data: Lista de dicionários para salvar
        """
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    
    def _generate_id(self, data: List[Dict[str, Any]]) -> int:
        """
        Gera um novo ID único baseado nos dados existentes.
        
        Args:
            data: Lista de dicionários com dados existentes
            
        Returns:
            Novo ID único
        """
        if not data:
            return 1
        return max(item.get('id', 0) for item in data) + 1
    
    # ========================================================================
    # OPERAÇÕES CRUD - FUNCIONÁRIOS (Nomenclatura em Português)
    # ========================================================================
    
    def adicionar_funcionario(self, employee_data: Dict[str, Any]) -> int:
        """
        Adiciona um novo funcionário ao banco de dados.
        
        Args:
            employee_data: Dicionário com dados do funcionário
            
        Returns:
            ID do funcionário criado
        """
        employees = self._load_json(self.employees_file)
        
        # Gerar novo ID
        new_id = self._generate_id(employees)
        employee_data['id'] = new_id
        
        # Adicionar timestamp de criação
        employee_data['created_at'] = datetime.now().isoformat()
        employee_data['updated_at'] = datetime.now().isoformat()
        
        # Adicionar à lista e salvar
        employees.append(employee_data)
        self._save_json(self.employees_file, employees)
        
        return new_id
    
    def buscar_funcionario(self, employee_id: int) -> Optional[Dict[str, Any]]:
        """
        Busca um funcionário por ID.
        
        Args:
            employee_id: ID do funcionário
            
        Returns:
            Dicionário com dados do funcionário ou None se não encontrado
        """
        employees = self._load_json(self.employees_file)
        
        for employee in employees:
            if employee.get('id') == employee_id:
                return employee
        
        return None
    
    def listar_funcionarios(self, apenas_ativos: bool = False) -> List[Dict[str, Any]]:
        """
        Retorna todos os funcionários cadastrados, com opção de filtrar apenas ativos.
        
        Args:
            apenas_ativos: Se True, retorna apenas funcionários com status 'Ativo'.
            
        Returns:
            Lista de dicionários com dados dos funcionários
        """
        if apenas_ativos:
            return self.buscar_funcionarios_por_status("Ativo")
        
        return self._load_json(self.employees_file)
    
    def atualizar_funcionario(self, employee_id: int, employee_data: Dict[str, Any]) -> bool:
        """
        Atualiza os dados de um funcionário.
        
        Args:
            employee_id: ID do funcionário
            employee_data: Dicionário com novos dados do funcionário
            
        Returns:
            True se atualizado com sucesso, False caso contrário
        """
        employees = self._load_json(self.employees_file)
        
        for i, employee in enumerate(employees):
            if employee.get('id') == employee_id:
                # Manter ID e timestamp de criação
                employee_data['id'] = employee_id
                employee_data['created_at'] = employee.get('created_at', datetime.now().isoformat())
                employee_data['updated_at'] = datetime.now().isoformat()
                
                # Atualizar funcionário
                employees[i] = employee_data
                self._save_json(self.employees_file, employees)
                return True
        
        return False
    
    def remover_funcionario(self, employee_id: int) -> bool:
        """
        Remove um funcionário do banco de dados.
        
        Args:
            employee_id: ID do funcionário
            
        Returns:
            True se removido com sucesso, False caso contrário
        """
        employees = self._load_json(self.employees_file)
        
        for i, employee in enumerate(employees):
            if employee.get('id') == employee_id:
                employees.pop(i)
                self._save_json(self.employees_file, employees)
                return True
        
        return False
    
    def buscar_funcionarios(self, query: str) -> List[Dict[str, Any]]:
        """
        Busca funcionários por nome ou matrícula.
        
        Args:
            query: Texto de busca
            
        Returns:
            Lista de funcionários que correspondem à busca
        """
        employees = self._load_json(self.employees_file)
        query_lower = query.lower()
        
        results = []
        for employee in employees:
            name = employee.get('name', '').lower()
            matricula = str(employee.get('matricula', '')).lower()
            
            if query_lower in name or query_lower in matricula:
                results.append(employee)
        
        return results
    
    def buscar_funcionarios_por_status(self, status: str) -> List[Dict[str, Any]]:
        """
        Retorna funcionários filtrados por status.
        
        Args:
            status: Status do funcionário (Ativo, Afastado, Desligado)
            
        Returns:
            Lista de funcionários com o status especificado
        """
        employees = self._load_json(self.employees_file)
        
        return [emp for emp in employees if emp.get('status') == status]
    
    # ========================================================================
    # OPERAÇÕES CRUD - AFASTAMENTOS (Nomenclatura em Português)
    # ========================================================================
    
    def adicionar_afastamento(self, absence_data: Dict[str, Any]) -> int:
        """
        Adiciona um novo afastamento ao banco de dados.
        
        Args:
            absence_data: Dicionário com dados do afastamento
            
        Returns:
            ID do afastamento criado
        """
        absences = self._load_json(self.afastamentos_file)
        
        # Gerar novo ID
        new_id = self._generate_id(absences)
        absence_data['id'] = new_id
        
        # Adicionar timestamp de criação
        absence_data['created_at'] = datetime.now().isoformat()
        absence_data['updated_at'] = datetime.now().isoformat()
        
        # Adicionar à lista e salvar
        absences.append(absence_data)
        self._save_json(self.afastamentos_file, absences)
        
        return new_id
    
    def buscar_afastamento(self, absence_id: int) -> Optional[Dict[str, Any]]:
        """
        Busca um afastamento por ID.
        
        Args:
            absence_id: ID do afastamento
            
        Returns:
            Dicionário com dados do afastamento ou None se não encontrado
        """
        absences = self._load_json(self.afastamentos_file)
        
        for absence in absences:
            if absence.get('id') == absence_id:
                return absence
        
        return None
    
    def listar_afastamentos(self) -> List[Dict[str, Any]]:
        """
        Retorna todos os afastamentos cadastrados.
        
        Returns:
            Lista de dicionários com dados dos afastamentos
        """
        return self._load_json(self.afastamentos_file)
    
    def buscar_afastamentos_por_funcionario(self, employee_id: int) -> List[Dict[str, Any]]:
        """
        Retorna todos os afastamentos de um funcionário específico.
        
        Args:
            employee_id: ID do funcionário
            
        Returns:
            Lista de afastamentos do funcionário
        """
        absences = self._load_json(self.afastamentos_file)
        
        return [abs for abs in absences if abs.get('employee_id') == employee_id]
    
    def atualizar_afastamento(self, absence_id: int, absence_data: Dict[str, Any]) -> bool:
        """
        Atualiza os dados de um afastamento.
        
        Args:
            absence_id: ID do afastamento
            absence_data: Dicionário com novos dados do afastamento
            
        Returns:
            True se atualizado com sucesso, False caso contrário
        """
        absences = self._load_json(self.afastamentos_file)
        
        for i, absence in enumerate(absences):
            if absence.get('id') == absence_id:
                # Manter ID e timestamp de criação
                absence_data['id'] = absence_id
                absence_data['created_at'] = absence.get('created_at', datetime.now().isoformat())
                absence_data['updated_at'] = datetime.now().isoformat()
                
                # Atualizar afastamento
                absences[i] = absence_data
                self._save_json(self.afastamentos_file, absences)
                return True
        
        return False
    
    def remover_afastamento(self, absence_id: int) -> bool:
        """
        Remove um afastamento do banco de dados.
        
        Args:
            absence_id: ID do afastamento
            
        Returns:
            True se removido com sucesso, False caso contrário
        """
        absences = self._load_json(self.afastamentos_file)
        
        for i, absence in enumerate(absences):
            if absence.get('id') == absence_id:
                absences.pop(i)
                self._save_json(self.afastamentos_file, absences)
                return True
        
        return False
    
    def listar_afastamentos_ativos(self) -> List[Dict[str, Any]]:
        """
        Retorna todos os afastamentos ativos (não finalizados).
        
        Returns:
            Lista de afastamentos ativos
        """
        absences = self._load_json(self.afastamentos_file)
        today = datetime.now().date()
        
        active = []
        for absence in absences:
            # Verificar se tem data de retorno
            if absence.get('return_date'):
                try:
                    return_date = datetime.fromisoformat(absence['return_date']).date()
                    if return_date >= today:
                        active.append(absence)
                except (ValueError, TypeError):
                    pass
            else:
                # Se não tem data de retorno, considera ativo
                active.append(absence)
        
        return active
    
    def buscar_afastamentos_por_periodo(self, start_date: str, end_date: str) -> List[Dict[str, Any]]:
        """
        Retorna afastamentos em um período específico.
        
        Args:
            start_date: Data inicial no formato ISO (YYYY-MM-DD)
            end_date: Data final no formato ISO (YYYY-MM-DD)
            
        Returns:
            Lista de afastamentos no período
        """
        absences = self._load_json(self.afastamentos_file)
        
        try:
            start = datetime.fromisoformat(start_date).date()
            end = datetime.fromisoformat(end_date).date()
        except (ValueError, TypeError):
            return []
        
        results = []
        for absence in absences:
            try:
                absence_start = datetime.fromisoformat(absence.get('start_date', '')).date()
                
                # Verificar se o afastamento está no período
                if absence.get('return_date'):
                    absence_end = datetime.fromisoformat(absence['return_date']).date()
                else:
                    absence_end = datetime.now().date()
                
                # Verificar sobreposição de períodos
                if absence_start <= end and absence_end >= start:
                    results.append(absence)
            except (ValueError, TypeError):
                continue
        
        return results
    
    # ========================================================================
    # OPERAÇÕES CRUD - USUÁRIOS (Nomenclatura em Português)
    # ========================================================================
    
    def adicionar_usuario(self, user_data: Dict[str, Any]) -> int:
        """
        Adiciona um novo usuário ao banco de dados.
        
        Args:
            user_data: Dicionário com dados do usuário
            
        Returns:
            ID do usuário criado
        """
        users = self._load_json(self.users_file)
        
        # Gerar novo ID
        new_id = self._generate_id(users)
        user_data['id'] = new_id
        
        # Adicionar timestamp de criação
        user_data['created_at'] = datetime.now().isoformat()
        user_data['updated_at'] = datetime.now().isoformat()
        
        # Adicionar à lista e salvar
        users.append(user_data)
        self._save_json(self.users_file, users)
        
        return new_id
    
    def buscar_usuario(self, user_id: int) -> Optional[Dict[str, Any]]:
        """
        Busca um usuário por ID.
        
        Args:
            user_id: ID do usuário
            
        Returns:
            Dicionário com dados do usuário ou None se não encontrado
        """
        users = self._load_json(self.users_file)
        
        for user in users:
            if user.get('id') == user_id:
                return user
        
        return None
    
    def buscar_usuario_por_username(self, username: str) -> Optional[Dict[str, Any]]:
        """
        Busca um usuário por nome de usuário.
        
        Args:
            username: Nome de usuário
            
        Returns:
            Dicionário com dados do usuário ou None se não encontrado
        """
        users = self._load_json(self.users_file)
        
        for user in users:
            if user.get('username') == username:
                return user
        
        return None
    
    def listar_usuarios(self) -> List[Dict[str, Any]]:
        """
        Retorna todos os usuários cadastrados.
        
        Returns:
            Lista de dicionários com dados dos usuários
        """
        return self._load_json(self.users_file)
    
    def atualizar_usuario(self, user_id: int, user_data: Dict[str, Any]) -> bool:
        """
        Atualiza os dados de um usuário.
        
        Args:
            user_id: ID do usuário
            user_data: Dicionário com novos dados do usuário
            
        Returns:
            True se atualizado com sucesso, False caso contrário
        """
        users = self._load_json(self.users_file)
        
        for i, user in enumerate(users):
            if user.get('id') == user_id:
                # Manter ID e timestamp de criação
                user_data['id'] = user_id
                user_data['created_at'] = user.get('created_at', datetime.now().isoformat())
                user_data['updated_at'] = datetime.now().isoformat()
                
                # Atualizar usuário
                users[i] = user_data
                self._save_json(self.users_file, users)
                return True
        
        return False
    
    def remover_usuario(self, user_id: int) -> bool:
        """
        Remove um usuário do banco de dados.
        
        Args:
            user_id: ID do usuário
            
        Returns:
            True se removido com sucesso, False caso contrário
        """
        users = self._load_json(self.users_file)
        
        for i, user in enumerate(users):
            if user.get('id') == user_id:
                users.pop(i)
                self._save_json(self.users_file, users)
                return True
        
        return False
    
    def validar_usuario(self, username: str, password: str) -> Optional[Dict[str, Any]]:
        """
        Valida credenciais de usuário.
        
        Args:
            username: Nome de usuário
            password: Senha do usuário
            
        Returns:
            Dicionário com dados do usuário se válido, None caso contrário
        """
        user = self.buscar_usuario_por_username(username)
        
        if user and user.get('password') == password:
            return user
        
        return None
    
    # ========================================================================
    # MÉTODOS DE EXPORTAÇÃO (Nomenclatura em Português)
    # ========================================================================
    
    def exportar_funcionarios_para_excel(self, filepath: str) -> bool:
        """
        Exporta dados de funcionários para arquivo Excel.
        
        Args:
            filepath: Caminho do arquivo Excel de destino
            
        Returns:
            True se exportado com sucesso, False caso contrário
        """
        try:
            employees = self.listar_funcionarios()
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Funcionários"
            
            # Definir cabeçalhos
            headers = [
                "ID", "Matrícula", "Nome", "CPF", "Cargo", 
                "Setor", "Status", "Data Admissão", "Telefone", "Email"
            ]
            
            # Estilizar cabeçalhos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Adicionar dados
            for row, employee in enumerate(employees, start=2):
                ws.cell(row=row, column=1, value=employee.get('id', ''))
                ws.cell(row=row, column=2, value=employee.get('matricula', ''))
                ws.cell(row=row, column=3, value=employee.get('name', ''))
                ws.cell(row=row, column=4, value=employee.get('cpf', ''))
                ws.cell(row=row, column=5, value=employee.get('cargo', ''))
                ws.cell(row=row, column=6, value=employee.get('setor', ''))
                ws.cell(row=row, column=7, value=employee.get('status', ''))
                ws.cell(row=row, column=8, value=employee.get('data_admissao', ''))
                ws.cell(row=row, column=9, value=employee.get('telefone', ''))
                ws.cell(row=row, column=10, value=employee.get('email', ''))
            
            # Ajustar largura das colunas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Salvar arquivo
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"Erro ao exportar funcionários: {e}")
            return False
    
    def exportar_afastamentos_para_excel(self, filepath: str) -> bool:
        """
        Exporta dados de afastamentos para arquivo Excel.
        
        Args:
            filepath: Caminho do arquivo Excel de destino
            
        Returns:
            True se exportado com sucesso, False caso contrário
        """
        try:
            absences = self.listar_afastamentos()
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Afastamentos"
            
            # Definir cabeçalhos
            headers = [
                "ID", "ID Funcionário", "Tipo", "Data Início", 
                "Data Retorno", "Dias", "Motivo", "Observações"
            ]
            
            # Estilizar cabeçalhos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Adicionar dados
            for row, absence in enumerate(absences, start=2):
                ws.cell(row=row, column=1, value=absence.get('id', ''))
                ws.cell(row=row, column=2, value=absence.get('employee_id', ''))
                ws.cell(row=row, column=3, value=absence.get('type', ''))
                ws.cell(row=row, column=4, value=absence.get('start_date', ''))
                ws.cell(row=row, column=5, value=absence.get('return_date', ''))
                ws.cell(row=row, column=6, value=absence.get('days', ''))
                ws.cell(row=row, column=7, value=absence.get('reason', ''))
                ws.cell(row=row, column=8, value=absence.get('observations', ''))
            
            # Ajustar largura das colunas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Salvar arquivo
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"Erro ao exportar afastamentos: {e}")
            return False
    
    # ========================================================================
    # MÉTODOS DE ESTATÍSTICAS E RELATÓRIOS (Nomenclatura em Português)
    # ========================================================================
    
    def obter_estatisticas(self) -> Dict[str, Any]:
        """
        Retorna estatísticas gerais do sistema.
        
        Returns:
            Dicionário com estatísticas do sistema
        """
        employees = self.listar_funcionarios()
        absences = self.listar_afastamentos()
        
        # Contar funcionários por status
        status_count = {}
        for employee in employees:
            status = employee.get('status', 'Desconhecido')
            status_count[status] = status_count.get(status, 0) + 1
        
        # Contar afastamentos por tipo
        absence_type_count = {}
        for absence in absences:
            absence_type = absence.get('type', 'Desconhecido')
            absence_type_count[absence_type] = absence_type_count.get(absence_type, 0) + 1
        
        # Contar afastamentos ativos
        active_absences = len(self.listar_afastamentos_ativos())
        
        return {
            'total_employees': len(employees),
            'employees_by_status': status_count,
            'total_absences': len(absences),
            'absences_by_type': absence_type_count,
            'active_absences': active_absences
        }
    
    def obter_funcionario_com_afastamentos(self, employee_id: int) -> Optional[Dict[str, Any]]:
        """
        Retorna dados completos de um funcionário incluindo seus afastamentos.
        
        Args:
            employee_id: ID do funcionário
            
        Returns:
            Dicionário com dados do funcionário e lista de afastamentos
        """
        employee = self.buscar_funcionario(employee_id)
        
        if not employee:
            return None
        
        absences = self.buscar_afastamentos_por_funcionario(employee_id)
        
        return {
            'employee': employee,
            'absences': absences
        }
    
    # ========================================================================
    # MÉTODOS DE BACKUP E MANUTENÇÃO (Nomenclatura em Português)
    # ========================================================================
    
    def fazer_backup(self, backup_dir: str) -> bool:
        """
        Cria backup de todos os arquivos de dados.
        
        Args:
            backup_dir: Diretório de destino para o backup
            
        Returns:
            True se backup criado com sucesso, False caso contrário
        """
        try:
            import shutil
            
            # Criar diretório de backup se não existir
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
            
            # Criar timestamp para o backup
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_subdir = os.path.join(backup_dir, f"backup_{timestamp}")
            os.makedirs(backup_subdir)
            
            # Copiar arquivos
            shutil.copy2(self.employees_file, backup_subdir)
            shutil.copy2(self.afastamentos_file, backup_subdir)
            shutil.copy2(self.users_file, backup_subdir)
            
            return True
            
        except Exception as e:
            print(f"Erro ao criar backup: {e}")
            return False
    
    def restaurar_backup(self, backup_dir: str) -> bool:
        """
        Restaura dados de um backup.
        
        Args:
            backup_dir: Diretório contendo os arquivos de backup
            
        Returns:
            True se restaurado com sucesso, False caso contrário
        """
        try:
            import shutil
            
            # Verificar se os arquivos de backup existem
            backup_employees = os.path.join(backup_dir, "employees.json")
            backup_absences = os.path.join(backup_dir, "absences.json")
            backup_users = os.path.join(backup_dir, "users.json")
            
            if not all(os.path.exists(f) for f in [backup_employees, backup_absences, backup_users]):
                return False
            
            # Restaurar arquivos
            shutil.copy2(backup_employees, self.employees_file)
            shutil.copy2(backup_absences, self.afastamentos_file)
            shutil.copy2(backup_users, self.users_file)
            
            return True
            
        except Exception as e:
            print(f"Erro ao restaurar backup: {e}")
            return False
    
    def limpar_todos_dados(self) -> bool:
        """
        Remove todos os dados do sistema (usar com cuidado).
        
        Returns:
            True se dados removidos com sucesso, False caso contrário
        """
        try:
            self._save_json(self.employees_file, [])
            self._save_json(self.afastamentos_file, [])
            self._save_json(self.users_file, [])
            return True
        except Exception as e:
            print(f"Erro ao limpar dados: {e}")
            return False


# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

def create_default_database() -> DatabaseManager:
    """
    Cria uma instância padrão do gerenciador de banco de dados.
    
    Returns:
        Instância configurada do DatabaseManager
    """
    return DatabaseManager()


def migrate_to_sql(json_db: DatabaseManager, sql_db) -> bool:
    """
    Migra dados do banco JSON para o banco SQL.
    
    Args:
        json_db: Instância do DatabaseManager (JSON)
        sql_db: Instância do DatabaseSQL
        
    Returns:
            True se migração bem-sucedida, False caso contrário
    """
    try:
        # Migrar funcionários
        employees = json_db.listar_funcionarios()
        for employee in employees:
            sql_db.add_employee(employee)
        
        # Migrar afastamentos
        absences = json_db.listar_afastamentos()
        for absence in absences:
            sql_db.add_absence(absence)
        
        # Migrar usuários
        users = json_db.listar_usuarios()
        for user in users:
            sql_db.add_user(user)
        
        return True
        
    except Exception as e:
            print(f"Erro na migração: {e}")
            return False


# ============================================================================
# PONTO DE ENTRADA PARA TESTES
# ============================================================================

if __name__ == "__main__":
    # Criar instância do banco de dados
    db = DatabaseManager()
    
    # Exemplo de uso
    print("=== Sistema RH Control - Database Manager (JSON) ===")
    print(f"Total de funcionários: {len(db.listar_funcionarios())}")
    print(f"Total de afastamentos: {len(db.listar_afastamentos())}")
    print(f"Total de usuários: {len(db.listar_usuarios())}")
    
    # Exibir estatísticas
    stats = db.obter_estatisticas()
    print("\n=== Estatísticas ===")
    print(f"Funcionários por status: {stats['employees_by_status']}")
    print(f"Afastamentos por tipo: {stats['absences_by_type']}")
    print(f"Afastamentos ativos: {stats['active_absences']}")
